"""
Device Spreadsheet Parser for HEMS
Handles WIDE/PIVOT FORMAT spreadsheets where:
- Each ROW = one room
- Each DEVICE TYPE = a column group (quantity + watt)
- Floor column uses merged cells (forward-fill)
- AC has extra columns (brand, star, ton, ISEER)

Uses openpyxl for merged cell handling.
"""

import pandas as pd
import openpyxl
import re
import json
from typing import List, Dict, Any, Optional
from django.conf import settings
from langchain_groq import ChatGroq
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.output_parsers import PydanticOutputParser
from pydantic import BaseModel, Field


class DeviceData(BaseModel):
    """Pydantic model for device data validation"""
    building: str = Field(default="Main Building")
    room: str = Field(default="General")
    device_type: str = Field(default="OTHER")
    sub_type: Optional[str] = Field(default=None)
    brand: str = Field(default="Generic")
    quantity: int = Field(ge=1, default=1)
    watt_rating: float = Field(gt=0, default=100.0)
    hours_used_per_day: float = Field(ge=0, le=24, default=0.0)
    star_rating: Optional[int] = Field(ge=1, le=5, default=None)
    ton: Optional[float] = Field(gt=0, default=None)



# Default watt ratings for when spreadsheet has no watt value
DEFAULT_WATTS = {
    "LIGHT": 20.0,
    "FAN": 75.0,
    "AC": 1500.0,
    "COOLER": 200.0,
    "PC": 400.0,
    "PROJECTOR": 300.0,
    "PRINTER": 500.0,
    "REFRIGERATOR": 200.0,
    "MICROWAVE": 1000.0,
    "INDUCTION": 2000.0,
    "TV": 150.0,
    "OTHER": 100.0,
}


class DeviceSpreadsheetParser:
    """
    Parser for WIDE/PIVOT FORMAT energy audit spreadsheets.
    
    Spreadsheet structure:
      Row = Floor | Building | Room | Device1_Qty | Device1_Watt | ...
    
    Unpivots into flat device list.
    """
    
    def __init__(self):
        """Initialize parser"""
        api_key = getattr(settings, 'GROQ_API_KEY', None)
        if api_key:
            self.llm = ChatGroq(
                groq_api_key=api_key,
                model_name="mixtral-8x7b-32768",
                temperature=0
            )
            self.output_parser = PydanticOutputParser(pydantic_object=DeviceData)
        else:
            self.llm = None
            self.output_parser = None
    
    def parse_spreadsheet(self, file_path: str, use_llm: bool = False) -> List[Dict[str, Any]]:
        """
        Main entry point. Auto-detects format.
        
        Args:
            file_path: Path to CSV or Excel file
            use_llm: Whether to use LLM (ignored for wide format)
            
        Returns:
            Flat list of device dicts (one per device per room)
        """
        if file_path.endswith(('.xlsx', '.xls')):
            devices = self._parse_wide_format(file_path)
            if devices:
                print(f"Wide-format parsing: {len(devices)} devices from {file_path}")
                return devices
            print("Wide-format returned 0 devices, trying simple format...")
        
        # Fallback: CSV or simple Excel
        return self._parse_simple_format(file_path, use_llm)
    
    # ========================================================================
    # WIDE FORMAT PARSER
    # ========================================================================
    
    def _parse_wide_format(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Parse wide/pivot format using exact column positions.
        """
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
        except Exception as e:
            print(f"openpyxl error: {e}")
            return []
        
        print(f"Sheet: {ws.title}, {ws.max_row} rows x {ws.max_column} cols")
        
        # Find header row (contains "ROOM" or "BUILDING")
        header_row = self._find_header_row(ws)
        print(f"Header row: {header_row}")
        
        # Dynamically detect column positions from header
        col_mapping = self._detect_columns(ws, header_row)
        
        if not col_mapping['devices']:
            print("No device columns detected!")
            return []
        
        print(f"Detected {len(col_mapping['devices'])} device column groups")
        print(f"Floor col: {col_mapping['floor_col']}, Building col: {col_mapping['building_col']}, Room col: {col_mapping['room_col']}")
        
        # Parse data rows
        devices = []
        current_floor = "Main Building"
        current_building = ""
        
        for row_idx in range(header_row + 1, ws.max_row + 1):
            # Floor (col 1) — forward-fill merged cells
            floor_val = ws.cell(row=row_idx, column=col_mapping['floor_col']).value
            if floor_val and str(floor_val).strip() and str(floor_val).strip() != ' ':
                floor_text = str(floor_val).strip()
                # Skip if looks like header repeat
                if not self._is_header_text(floor_text):
                    current_floor = floor_text
            
            # Building/dept (col 2)
            bldg_val = ws.cell(row=row_idx, column=col_mapping['building_col']).value
            if bldg_val and str(bldg_val).strip():
                bldg_text = str(bldg_val).strip()
                if not self._is_header_text(bldg_text):
                    current_building = bldg_text
            
            # Room (col 3)
            room_val = ws.cell(row=row_idx, column=col_mapping['room_col']).value
            if not room_val:
                continue
            
            room_str = str(room_val).strip()
            
            # Skip header/label rows
            if self._is_header_text(room_str):
                continue
            if not room_str or room_str.upper() == 'NAN':
                continue
            
            # Build display building name
            building_name = current_floor
            if current_building:
                building_name = f"{current_floor} - {current_building}"
            
            # Extract standard devices
            for qty_col, watt_col, device_type, sub_type in col_mapping['devices']:
                qty = self._safe_int(ws.cell(row=row_idx, column=qty_col).value)
                if qty <= 0:
                    continue
                
                watt = self._safe_float(ws.cell(row=row_idx, column=watt_col).value)
                if watt <= 0:
                    watt = DEFAULT_WATTS.get(device_type, 100.0)
                
                devices.append({
                    'building': building_name,
                    'room': room_str,
                    'device_type': device_type,
                    'sub_type': sub_type,
                    'brand': 'Generic',
                    'quantity': qty,
                    'watt_rating': watt,
                    'hours_used_per_day': 0.0,
                    'star_rating': None,
                    'ton': None,
                })
            
            # Extract AC (special columns)
            if col_mapping.get('ac'):
                ac = col_mapping['ac']
                ac_qty = self._safe_int(ws.cell(row=row_idx, column=ac['qty_col']).value)
                
                if ac_qty > 0:
                    ac_watt = self._safe_float(ws.cell(row=row_idx, column=ac['watt_col']).value)
                    if ac_watt <= 0:
                        ac_watt = 1500.0
                    
                    # Brand
                    brand_val = ws.cell(row=row_idx, column=ac['brand_col']).value
                    brand = 'Generic'
                    if brand_val and str(brand_val).strip():
                        b = str(brand_val).strip()
                        if b.upper() not in ('0', 'NAN', 'NONE', ''):
                            brand = b
                    
                    # Star rating
                    star = self._safe_int(ws.cell(row=row_idx, column=ac['star_col']).value)
                    star = star if 1 <= star <= 5 else None
                    
                    # Tonnage
                    ton = self._safe_float(ws.cell(row=row_idx, column=ac['ton_col']).value)
                    ton = ton if ton > 0 else None
                    
                    devices.append({
                        'building': building_name,
                        'room': room_str,
                        'device_type': 'AC',
                        'sub_type': 'Split AC',
                        'brand': brand,
                        'quantity': ac_qty,
                        'watt_rating': ac_watt,
                        'hours_used_per_day': 0.0,
                        'star_rating': star,
                        'ton': ton,
                    })
        
        return devices
    
    def _find_header_row(self, ws) -> int:
        """Find the header row by scanning for ROOM or BUILDING text"""
        for row_idx in range(1, min(11, ws.max_row + 1)):
            for col_idx in range(1, min(10, ws.max_column + 1)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val and str(val).strip().upper() in ('ROOM', 'ROOM ', 'ROOM NO', 'BUILDING'):
                    return row_idx
        return 1  # Default to row 1
    
    def _detect_columns(self, ws, header_row: int) -> Dict:
        """
        Dynamically detect column positions from header row.
        Returns dict with floor_col, building_col, room_col, devices list, ac dict.
        """
        result = {
            'floor_col': 1,
            'building_col': 2, 
            'room_col': 3,
            'devices': [],
            'ac': None,
        }
        
        # Read all headers
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col_idx).value
            if val:
                headers[col_idx] = str(val).strip().upper()
        
        print(f"Headers: {headers}")
        
        # Find floor/building/room columns
        for col_idx, val in headers.items():
            if val in ('BUILDING', 'BUILDING '):
                result['building_col'] = col_idx
            elif val in ('ROOM', 'ROOM ', 'ROOM NO', 'ROOM NO.'):
                result['room_col'] = col_idx
        
        # Device keyword mapping: header text → (device_type, sub_type)
        device_patterns = {
            '2X2': ('LIGHT', '2x2 Panel'),
            'LED-TB': ('LIGHT', 'LED Tube'),
            'LED TB': ('LIGHT', 'LED Tube'),
            'SENSOR': ('LIGHT', 'Sensor Light'),
            'ROUND': ('LIGHT', 'Round Tube'),
            'FLOODE': ('LIGHT', 'Flood Light'),
            'FLOOD': ('LIGHT', 'Flood Light'),
            'SQUARE': ('LIGHT', 'Square Light'),
            'CIRCLE': ('LIGHT', 'Circle Light'),
            'C FAN': ('FAN', 'Ceiling Fan'),
            'CEILING FAN': ('FAN', 'Ceiling Fan'),
            'WALL FAN': ('FAN', 'Wall Fan'),
            'TABLE FAN': ('FAN', 'Table Fan'),
            'EXHAUST': ('FAN', 'Exhaust Fan'),
            'COOLER': ('COOLER', 'Desert Cooler'),
            'PROJECTOR': ('PROJECTOR', 'Projector'),
            'PC': ('PC', 'Desktop'),
            'PRINTER': ('PRINTER', 'Printer'),
            'RF': ('REFRIGERATOR', 'Refrigerator'),
            'REFRIGERATOR': ('REFRIGERATOR', 'Refrigerator'),
            'MICROWAVE': ('MICROWAVE', 'Microwave'),
            'INDUCTION': ('INDUCTION', 'Induction'),
            'TV': ('TV', 'TV'),
        }
        
        # Scan headers for device columns
        processed_cols = set()
        
        for col_idx, val in sorted(headers.items()):
            if col_idx in processed_cols:
                continue
            
            # Check for AC special structure
            if val in ('AC COMPANY', 'AC COMP'):
                # AC has: BRAND(col), QTY(col+1), STAR(col+2), WATT(col+3), TON(col+4), ISEER(col+5)
                result['ac'] = {
                    'brand_col': col_idx,
                    'qty_col': col_idx + 1,
                    'star_col': col_idx + 2,
                    'watt_col': col_idx + 3,
                    'ton_col': col_idx + 4,
                    'iseer_col': col_idx + 5,
                }
                for i in range(6):
                    processed_cols.add(col_idx + i)
                print(f"  AC detected: brand={col_idx}, qty={col_idx+1}, star={col_idx+2}, watt={col_idx+3}, ton={col_idx+4}")
                continue
            
            # Skip WATT columns (they belong to previous device)
            if val == 'WATT':
                processed_cols.add(col_idx)
                continue
            
            # Try to match device pattern
            for pattern, (dev_type, sub_type) in device_patterns.items():
                if pattern in val:
                    # This col = qty, next col = watt
                    watt_col = col_idx + 1
                    result['devices'].append((col_idx, watt_col, dev_type, sub_type))
                    processed_cols.add(col_idx)
                    processed_cols.add(watt_col)
                    print(f"  {val} -> {dev_type}/{sub_type} (qty={col_idx}, watt={watt_col})")
                    break
        
        return result
    
    def _is_header_text(self, text: str) -> bool:
        """Check if text looks like a header row (not data)"""
        upper = text.upper().strip()
        return upper in (
            'BUILDING', 'ROOM', 'ROOM NO', 'ROOM NO.', 'SR', 'SR NO', 'SR.',
            'FLOOR', 'BLOCK', 'LIGHT', 'FAN', 'AC', 'WATT',
            'LIGHT ', 'FAN ', 'BUILDING ',
        )
    
    # ========================================================================
    # SIMPLE FORMAT FALLBACK
    # ========================================================================
    
    def _parse_simple_format(self, file_path: str, use_llm: bool = False) -> List[Dict[str, Any]]:
        """Fallback for CSV / simple one-device-per-row files"""
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
        
        if df.empty:
            return []
        
        df.columns = [str(c).strip() for c in df.columns]
        devices = []
        
        for idx, row in df.iterrows():
            row_dict = row.to_dict()
            
            if use_llm and self.llm:
                device = self._parse_row_with_llm(row_dict)
            else:
                device = self._parse_row_heuristic(row_dict)
            
            if device:
                devices.append(device)
        
        return devices
    
    def _parse_row_heuristic(self, row_data: Dict[str, Any]) -> Dict[str, Any]:
        """Simple heuristic parsing for one-device-per-row format"""
        device = {
            'building': 'Main Building', 'room': 'General',
            'device_type': 'OTHER', 'sub_type': None,
            'brand': 'Generic', 'quantity': 1,
            'watt_rating': 100.0, 'hours_used_per_day': 0.0,
            'star_rating': None, 'ton': None,
        }
        row_lower = {k.lower(): v for k, v in row_data.items()}
        
        for key in ['location', 'building', 'block', 'floor']:
            if key in row_lower and pd.notna(row_lower[key]):
                device['building'] = str(row_lower[key]).strip()
                break
        
        for key in ['area', 'room', 'room no']:
            if key in row_lower and pd.notna(row_lower[key]):
                device['room'] = str(row_lower[key]).strip()
                break
        
        for key in ['brand', 'make']:
            if key in row_lower and pd.notna(row_lower[key]):
                device['brand'] = str(row_lower[key]).strip()
                break
        
        for key in ['qty', 'quantity', 'count']:
            if key in row_lower and pd.notna(row_lower[key]):
                qty = self._safe_int(row_lower[key])
                if qty > 0: device['quantity'] = qty
                break
        
        for key in ['power', 'watt', 'watts', 'watt_rating']:
            if key in row_lower and pd.notna(row_lower[key]):
                watt = self._safe_float(row_lower[key])
                if watt > 0: device['watt_rating'] = watt
                break
        
        return device
    
    def _parse_row_with_llm(self, row_data: Dict[str, Any]) -> Dict[str, Any]:
        """Parse a single row using LLM"""
        flat_text = " | ".join(str(v) for v in row_data.values() if pd.notna(v))
        
        prompt = ChatPromptTemplate.from_template(
            """You are a strict electrical audit parser.
Extract structured device data from this row.
Rules:
1. Detect device type (FAN, AC, LIGHT, PC, COOLER, PROJECTOR, PRINTER, TV, OTHER).
2. Detect subtype.
3. Extract quantity (default 1).
4. Extract watt rating (convert kW to W). If missing, estimate realistic watt.
5. Extract brand.
Return strict JSON.

{format_instructions}

Row: {row_text}
"""
        )
        chain = prompt | self.llm | self.output_parser
        try:
            result = chain.invoke({
                "row_text": flat_text,
                "format_instructions": self.output_parser.get_format_instructions()
            })
            return result.dict()
        except:
            return self._parse_row_heuristic(row_data)
    
    # ========================================================================
    # UTILITIES
    # ========================================================================
    
    def _safe_int(self, val) -> int:
        if val is None: return 0
        try: return int(float(str(val).strip()))
        except: return 0
    
    def _safe_float(self, val) -> float:
        if val is None: return 0.0
        try:
            s = str(val).strip().upper()
            if 'KW' in s: return float(s.replace('KW', '').strip()) * 1000
            s = s.replace('W', '').replace('WATT', '').strip()
            return float(s)
        except: return 0.0
    
    def validate_devices(self, devices: List[Dict[str, Any]]) -> tuple:
        valid, errors = [], []
        for idx, d in enumerate(devices):
            try:
                DeviceData(**d)
                valid.append(d)
            except Exception as e:
                errors.append(f"Row {idx+1}: {str(e)}")
        return valid, errors
